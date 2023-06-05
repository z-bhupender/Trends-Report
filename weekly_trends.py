import boto3
import smart_open
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import openpyxl.styles as styles


class AWSHandler:
    def __init__(self) -> None:
        self.athena = boto3.client("athena", region_name="us-west-2")

    def run_query(self, query: str) -> pd.DataFrame:
        response = self.athena.start_query_execution(
            QueryString=query,
            QueryExecutionContext={"Database": "prodcoach"},
            ResultConfiguration={
                "OutputLocation": "s3://z-ally-proc-oregon-greenzone/query_results_new/athena/output/"
            },
        )

        execution_id = response["QueryExecutionId"]

        while True:
            query_execution = self.athena.get_query_execution(
                QueryExecutionId=execution_id
            )
            state = query_execution["QueryExecution"]["Status"]["State"]

            if state == "SUCCEEDED":
                break
            elif state == "FAILED" or state == "CANCELLED":
                raise Exception(f"Athena query {state.lower()}.")

        result_location = query_execution["QueryExecution"]["ResultConfiguration"][
            "OutputLocation"
        ]

        file_content = pd.read_csv(smart_open.open(result_location))

        if len(file_content) > 0:
            return file_content
        else:
            raise Exception(
                "No Calls Are Present For The Agent Id's In This Date Range!"
            )


class TrendsReport(object):
    def __init__(self, start_date: str, end_date: str, agent_id: str) -> None:
        handler = AWSHandler()
        where_clause = f"a.timestamp >= '{start_date}T00:00:00' AND a.timestamp <= '{end_date}T23:59:59' and zid1.internal_id in ( {agent_id} )"
        query = f'SELECT zid1.internal_id as agent_id, zid1.agent_name as agent_name, zid2.id as supervisor_id, zid2.agent_name as supervisor_name, cti.callid._id AS call_id, a.scores."rules.open.confirm_customer_id".AI_Scores.score AS "rules.open.confirm_customer_id.score", a.scores."rules.open.confirm_agent_id".AI_Scores.score AS "rules.open.confirm_agent_id.score", a.scores."rules.open.develop_rpc".AI_Scores.score AS "rules.open.develop_rpc.score", a.scores."rules.open.state_mini_miranda".AI_Scores.score AS "rules.open.state_mini_miranda.score", a.scores."rules.open.state_call_monitor".AI_Scores.score AS "rules.open.outbound.state_call_monitor.score", a.scores."rules.open.outbound.call_purpose".AI_Scores.score AS "rules.open.outbound.call_purpose.score", a.scores."rule.open.pause_listen".AI_Scores.score AS "rules.open.pause_listen.score", a.scores."rules.open.inbound.account_id".AI_Scores.score AS "rules.open.inbound.account_id.score", a.scores."rules.open.inbound.verify_identity".AI_Scores.score AS "rules.open.inbound.verify_identity.score", a.scores."rules.facts_on_the_table.pay_total_amount_due_today".AI_Scores.score AS "rules.facts_on_the_table.pay_total_amount_due_today.score", a.scores."rules.dqs.how_much_pay_today".AI_Scores.score AS "rules.dqs.how_much_pay_today.score", a.scores."rules.dqs.when_pay_remaining".AI_Scores.score AS "rules.dqs.when_pay_remaining.score", a.scores."rules.negotiation_flow.paying.gain_tad".AI_Scores.score AS "rules.negotiation_flow.paying.gain_tad.score", a.scores."rules.negotiation_flow.rapid_payment".AI_Scores.score AS "rules.negotiation_flow.rapid_payment.score", a.scores."rules.negotiation_flow.paying.straight_to_close".AI_Scores.score AS "rules.negotiation_flow.paying.straight_to_close.score", a.scores."rules.negotiation_flow.willing.two.payments".AI_Scores.score AS "rules.negotiation_flow.willing.two.payments.score", a.scores."rules.negotiation_flow.willing.raise_offer".AI_Scores.score AS "rules.negotiation_flow.willing.raise_offer.score", a.scores."rules.negotiation_flow.vague.partial_vague".AI_Scores.score AS "rules.negotiation_flow.vague.partial_vague.score", a.scores."rules.negotiation_flow.probing_questions".AI_Scores.score AS "rules.negotiation_flow.probing_questions.score", a.scores."rules.negotiation_flow.vague.3questions".AI_Scores.score AS "rules.negotiation_flow.vague.3questions.score", a.scores."rules.negotiation_flow.vague.repeat_until_full".AI_Scores.score AS "rules.negotiation_flow.vague.repeat_until_full.score", a.scores."rules.negotiation_flow.other_income".AI_Scores.score AS "rules.negotiation_flow.other_income.score", a.scores."rules.negotiation_flow.unwilling.rfd".AI_Scores.score AS "rules.negotiation_flow.unwilling.rfd.score", a.scores."rules.negotiation_flow.unwilling.sources_income".AI_Scores.score AS "rules.negotiation_flow.unwilling.sources_income.score", a.scores."rules.negotiation_flow.unwilling.create_solution".AI_Scores.score AS "rules.negotiation_flow.unwilling.create_solution.score", a.scores."rules.negotiation_flow.unwilling.modification".AI_Scores.score AS "rules.negotiation_flow.unwilling.modification.score", a.scores."rules.negotiation_flow.unwilling.intent".AI_Scores.score AS "rules.negotiation_flow.unwilling.intent.score", a.scores."rules.close.demographics".AI_Scores.score AS "rules.close.demographics.score", a.scores."rules.close.urgency".AI_Scores.score AS "rules.close.urgency.score", a.scores."rules.close.partial.remainder".AI_Scores.score AS "rules.close.partial.remainder.score", a.scores."rules.close.recap".AI_Scores.score AS "rules.close.recap.score", a.scores."rules.close.thank".AI_Scores.score AS "rules.close.thank.score", a.scores."rules.close.none.mission".AI_Scores.score AS "rules.close.none.mission.score", a.scores."rules.emotional_outburst.aet.acknowledge".AI_Scores.score AS "rules.emotional_outburst.aet.acknowledge.score", a.scores."rules.emotional_outburst.aet.wiifm".AI_Scores.score AS "rules.emotional_outburst.aet.wiifm.score", a.scores."rules.emotional_outburst.bridge.active_listening".AI_Scores.score AS "rules.emotional_outburst.bridge.active_listening.score", a.scores."rules.emotional_outburst.bridge.acknowledge_emotion".AI_Scores.score AS "rules.emotional_outburst.bridge.acknowledge_emotion.score", a.scores."rules.emotional_outburst.bridge.remove_isolation".AI_Scores.score AS "rules.emotional_outburst.bridge.remove_isolation.score", a.scores."rules.emotional_outburst.bridge.assure_customer".AI_Scores.score AS "rules.emotional_outburst.bridge.assure_customer.score" FROM metadata_datastore_new mnew JOIN coach_scorecards_ai_scores_luis a ON cast(cti.callid._id AS varchar) = a.call_id JOIN zid_new_hudi zid1 on cti.contact._core_agentpk = zid1.internal_id JOIN zid_new_hudi zid2 on zid1.supervisor_id = zid2.id where {where_clause}'
        print(f"Query ----->\n{query}")
        self.data = handler.run_query(query)
        print(f"\nAthena Result ----->\n{self.data}")
        self.start_date = start_date
        self.end_date = end_date
        self.load_file = self.data
        self.agent_data = self.load_file.iloc[:, [0, 1, 2, 3, 4]].copy()
        self.scores_data = self.load_file.iloc[
            :, [0, *[i for i in range(4, 44)]]
        ].copy()

    def get_call_id_per_agent_id(self, df: pd.DataFrame) -> dict:
        data = {}

        for row in df.iterrows():
            try:
                data[row[1]["agent_id"]].append(row[1]["call_id"])
            except KeyError:
                data[row[1]["agent_id"]] = [row[1]["call_id"]]

        return data

    def get_agent_name_per_agent_id(self, df: pd.DataFrame) -> dict:
        return dict(zip(df["agent_id"], df["agent_name"]))

    def get_supervisor_name_per_agent_id(self, df: pd.DataFrame) -> dict:
        return dict(zip(df["agent_id"], df["supervisor_name"]))

    def get_supervisor_id_per_agent_id(self, df: pd.DataFrame) -> dict:
        return dict(zip(df["agent_id"], df["supervisor_id"]))

    def get_skill_name_and_score_per_agent_id(self) -> dict:
        data = self.get_call_id_per_agent_id(self.agent_data)
        return {
            agent: self.select_skills(
                [(val, agent) for val in data[agent]], self.scores_data
            )
            for agent in data
        }

    def get_skill_list_by_agent(self) -> pd.DataFrame:
        skill_list_by_agent = self.get_skill_name_and_score_per_agent_id()
        agent_name = self.get_agent_name_per_agent_id(self.agent_data)
        supervisor_name = self.get_supervisor_name_per_agent_id(self.agent_data)
        supervisor_id = self.get_supervisor_id_per_agent_id(self.agent_data)
        return pd.DataFrame(
            [
                {
                    **skill,
                    "agent_id": agent,
                    "agent_name": agent_name[agent],
                    "supervisor_id": supervisor_id[agent],
                    "supervisor_name": supervisor_name[agent],
                }
                for agent in skill_list_by_agent
                for skill in skill_list_by_agent[agent]
            ]
        )

    def get_all_agents_per_skill(self, df: pd.DataFrame) -> dict:
        skill_list_by_agent = self.get_skill_list_by_agent()
        agent_by_skill = {
            skill: sorted(
                [
                    {
                        "agent_pk": row["agent_id"],
                        "agent_name": row["agent_name"],
                        "supervisor_id": row["supervisor_id"],
                        "supervisor_name": row["supervisor_name"],
                        "skill_score": row["skill_score"],
                        "call_count": row["call_count"],
                    }
                    for _, row in skill_list_by_agent[
                        skill_list_by_agent["skill_name"] == skill
                    ].iterrows()
                ],
                key=lambda d: d["skill_score"],
            )
            for skill in set(df.columns)
            & set(skill_list_by_agent["skill_name"].unique())
        }
        return agent_by_skill

    #! Change the val == 1 to val == "1" if query is from athena
    def get_yes(self, val: int) -> int:
        return 1 if val == 1 else 0

    def get_total_call(self, val: int):
        return 1 if val in [1, 2, 3] else 0

    def select_skills(self, call_list, df: pd.DataFrame):
        per_skill_calls = []
        selected_skills = []

        agent_df = df[df["agent_id"] == call_list[0][1]].drop_duplicates()
        per_skill_calls += [dict(row[1]) for row in agent_df.iterrows()]

        new_df = pd.DataFrame(per_skill_calls)

        for col in df.columns:
            if col not in ["call_id", "agent_id"]:
                temp_val = (
                    sum(new_df[col].apply(self.get_yes))
                    / sum(new_df[col].apply(self.get_total_call))
                    if sum(new_df[col].apply(self.get_total_call)) > 0
                    else 0
                )
                temp_val = round(temp_val * 2, 1)

                try:
                    selected_skills.append(
                        {
                            "skill_name": col,
                            "skill_score": temp_val,
                            "call_count": sum(new_df[col].apply(self.get_total_call)),
                        }
                    )
                except KeyError:
                    selected_skills = [
                        {
                            "skill_name": col,
                            "skill_score": temp_val,
                            "call_count": sum(new_df[col].apply(self.get_total_call)),
                        }
                    ]

        return sorted(selected_skills, key=lambda d: d["skill_score"])

    def get_dataFrame(self) -> pd.DataFrame:
        """
        /*
         * Returns a pandas dataframe in a new format
         * obtained from loaded json file.
         */
        """
        json_data = self.get_all_agents_per_skill(self.scores_data)
        return pd.DataFrame.from_dict(
            [
                {
                    "agent_pk": item["agent_pk"],
                    "agent_name": item["agent_name"],
                    "supervisor_id": item["supervisor_id"],
                    "supervisor_name": item["supervisor_name"],
                    "skill_details": {
                        "skill_name": skill_name,
                        "skill_score": item["skill_score"],
                        "call_count": item["call_count"],
                    },
                }
                for skill_name, value in json_data.items()
                for item in value
            ]
        )

    def get_agent_report(self) -> list:
        df = self.get_dataFrame()
        return [
            {
                "agent_pk": int(id),
                "agent_name": df.loc[df["agent_pk"] == id, "agent_name"].iloc[0],
                "supervisor_id": df.loc[df["agent_pk"] == id, "supervisor_id"].iloc[0],
                "supervisor_name": df.loc[df["agent_pk"] == id, "supervisor_name"].iloc[
                    0
                ],
                "skill_details": sorted(
                    df.loc[df["agent_pk"] == id, "skill_details"].tolist(),
                    key=lambda d: d["skill_score"],
                ),
            }
            for id in df["agent_pk"].unique()
        ]

    def run(self) -> list:
        return self.get_agent_report()


class ConverResultToExcel(object):
    def __init__(self) -> None:
        self.first_file_start_date = input("Start Date : ")
        self.first_file_end_date = input("End Date : ")
        get_agent_id = input("Enter Agent Id's Separated With Comma : ")
        self.agent_id = get_agent_id

        trends = TrendsReport(
            self.first_file_start_date, self.first_file_end_date, self.agent_id
        )
        get_json = trends.run()
        print(f"\nProcessed Data ----->\n{get_json}")
        self.df1 = get_json

        self.yellow_fill = self.get_Fill("FFF6BD")
        self.red_fill = self.get_Fill("FF9F9F")
        self.green_fill = self.get_Fill("CEEDC7")

        self.rule_desc = self.load_rules()

    def load_rules(self) -> list:
        return [
            {
                "ruleId": "rules.open.confirm_customer_id.score",
                "display_name": "ID Customer by First and Last Name (Including Suffix)",
                "rule_description": "Agent must obtain from the customer BOTH first and last name.  This may be in a single statement from the customer or multiple statements",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.confirm_agent_id.score",
                "display_name": "ID Self by First and Last Name and Company",
                "rule_description": "Agent must open the call by stating their first and last name including Ally Financial",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.develop_rpc.score",
                "display_name": "Develop RPC Info",
                "rule_description": "Agent must gain rpc info if customer not available (number, work, time available, etc.)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.outbound.state_call_monitor.score",
                "display_name": "Provide Recording Disclosure",
                "rule_description": 'Agent must state that the call may be "monitored" and/or "recorded" for quality assurance purposes.',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.state_mini_miranda.score",
                "display_name": "Provide Mini-Miranda (State Specifics)",
                "rule_description": 'If customer\'s address is in certain states, agent must include the statement "This is an attempt to collect a debt and any information obtained will be used for that purpose."',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.outbound.call_purpose.score",
                "display_name": "State the Purpose of your Call",
                "rule_description": "Agent must provide the Year, Make and Model of the vehicle they are calling the customer about",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.pause_listen.score",
                "display_name": "Pause / Listen",
                "description": "If customer makes an emotional statement, the agent should acknowledge the emotion and statement. They should not include any questions or attempts to hit the next step",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.inbound.account_id.score",
                "display_name": "Ask for Account Number or SS#",
                "rule_description": "Agent must obtain EITHER the account number or SSN to locate the account.  They can ask for both or just one.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.inbound.verify_identity.score",
                "display_name": "Ask for the Customer ID",
                "rule_description": "Agent must obtain 2 pieces of verification information.  This can be any of the following: Date of Birth, Last 4 of SSN (if SSN NOT used to pull up the account in item 9), Address,",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.facts_on_the_table.pay_total_amount_due_today.score",
                "display_name": "Facts on the table: You have a total amount due of ___. Can you pay that today?",
                "rule_description": 'This must be stated verbatim.  Agent should include "total amount due <$amount> " and "pay" "today"',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.dqs.how_much_pay_today.score",
                "display_name": "Diagnostic: How much can you pay today?",
                "rule_description": "Applicable if the answer to Facts on the Table is NO.  Question should be asked before asking when the customer can make a payment.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.dqs.when_pay_remaining.score",
                "display_name": "When can you pay the remaining balance of $___?",
                "rule_description": "Applicable if the answer to Facts on the Table is NO.  Question should be asked after asking how much the customer can pay today and BEFORE the agent asks additional questions.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.paying.gain_tad.score",
                "display_name": "Gained Initial Balance Commitment",
                "rule_description": "Agent obtains commitment for full balance in 2 installments prior to next due date.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.rapid_payment.score",
                "display_name": "Sell Rapid Payment",
                "rule_description": "Agent asks customer if they want to set up payment online and provides them benefits such as: easy, instant, free processs",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.paying.straight_to_close.score",
                "display_name": "Proceed Straight to Close",
                "rule_description": "Agent does not ask any probing questions regarding income or reason for delinquency and move to skills - Close Solution",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.willing.two.payments.score",
                "display_name": "Gained Initial Balance Commitment",
                "rule_description": "Agent obtains commitment for full balance in 2 installments prior to next due date.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.willing.raise_offer.score",
                "display_name": "Raised Initial Payment Offer",
                "rule_description": "Agent provided a benefit increasing the initial commitment made by the customer",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.vague.partial_vague.score",
                "display_name": "Agreed for Partial/Initial Payment",
                "rule_description": "Agent obtains commitment for 1 payment but customer would not commit to second payment",
                "sectionOffsetStart": "00:39.4",
                "sectionOffsetEnd": "00:57.8",
            },
            {
                "ruleId": "rules.negotiation_flow.probing_questions.score",
                "display_name": "Set the Stage",
                "rule_description": "Agents ask for permission to ask additional questions to further understand the customer's situation ",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.vague.3questions.score",
                "display_name": "Determined Next Income Period & Gained Commitment for Balance or Some Part",
                "rule_description": "Agent asks for when customer will be receiving next income. Agent asks for full remaining amount due on next income date. Agent asks how much customer can pay on next income date (if applicable)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.vague.repeat_until_full.score",
                "display_name": "Attempted to Continue Using Income Until Full Due is Committed/?",
                "rule_description": "Agent repeats income questions regarding pay periods and gaining partial commitments until total amount is committed",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.other_income.score",
                "display_name": "Determined Other Income in the Household",
                "rule_description": "Agent asks if customer is receiving severance (if applicable) and/or if there are alternate sources of income such as savings, 401k, other assets and/or the potential to borrow money from a friend or family member. Agent asks if customer can borrow money from a friend or family member",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.rfd.score",
                "display_name": "Did we ask for the Reason for Delinquency",
                "rule_description": "Agent asks for the reason the account fell behind ",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.sources_income.score",
                "display_name": "Asked for Regular Sources of Income",
                "rule_description": "Agent asks customer about their regular sources of income such as from a job or unemployment.  Agent asks about the frequency of income received.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.create_solution.score",
                "display_name": "Create & Sell Solutions Based on Sources Found",
                "rule_description": "Agent provides solutions such as borrowing money from the friends or family and paying back",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.modification.score",
                "display_name": "PROGRAM THAT FITS",
                "rule_description": "Offered Properly? (Extension, Modification, & Due Date Change)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.intent.score",
                "display_name": "Determined Customer's Intent (If Applicable / No Solution)",
                "rule_description": "Agents clarifies plan for payment in the future.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.demographics.score",
                "display_name": "Update Demographics",
                "rule_description": "Agent asks for ALL of the following: Customer's address, Garaging location of the vehicle, Best contact phone number, Employer name (if applicable), Employer phone number (if applicable). Agent should not provide any information on file, agent should ask the customer to provide the information.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.urgency.score",
                "display_name": "Create Urgency",
                "rule_description": "State importance of following through with plan for obtaining funds AND recaps negative result of not following through (to avoid additional fees) OR positive result (to get your account current again)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.partial.remainder.score",
                "display_name": "Create a Mission for the Remainder",
                "rule_description": 'Agent recaps customer\'s plan to obtain remainder of payment and commitment to call back with result.  Recap must include plan/action committed to (ie "talk to your brother about borrowing <remaining amount>") AND date to call back AND time to call back',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.recap.score",
                "display_name": "Recap Expectations of Customer",
                "rule_description": "If in repo status-inform customer car is out for repo/month end consequences, if applicable based on status of account",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.thank.score",
                "display_name": "Thanked the Customer for their Time",
                "rule_description": 'Agent expresses gratitude by "thank you" OR "glad" and wishes good day',
                "sectionOffsetStart": "03:08.3",
                "sectionOffsetEnd": "03:17.2",
            },
            {
                "ruleId": "rules.close.none.mission.score",
                "display_name": "Create a Mission",
                "rule_description": 'Agent recaps the customer\'s plan to obtain funds for payment and their commitment to call back with result.  Recap should include plan/action (ie "talk to your brother about borrowing <amount>") AND date to call back AND time to call back',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.aet.acknowledge.score",
                "display_name": "Acknowledged What was Said",
                "rule_description": "If customer makes an emotional statement, the agent should acknowledge the emotion and statement.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.aet.wiifm.score",
                "display_name": "Transitioned with a WIIFM",
                "rule_description": "Agent expresses desire to assist customer and provides a benefit to the customer for making a payment and/or staying on the line with the agent.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.active_listening.score",
                "display_name": "Active Listening",
                "rule_description": "Agent reflects backs information provided by the customer.  ",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.acknowledge_emotion.score",
                "display_name": "Acknowledged Customer's Emotion",
                "rule_description": "If customer makes an emotional statement, the agent should acknowledge the emotion and statement.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.remove_isolation.score",
                "display_name": "Did the Agent Attempt to Remove the Isolation",
                "rule_description": "Agent expresses desire to help, that they are the right person to assist the customer, lets the customer know there are many others who have experienced what is happening to the customer",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.assure_customer.score",
                "display_name": "Transitioned with Assuring the Customer that we can Find a Solution",
                "rule_description": "Agent expresses confidence and asks/tell the customer they wish to work together to find a solution and/or see what options are available.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
        ]

    def date_converter(self, p: str, q: str) -> str:
        a = p
        b = q
        x = a[5:7]
        y = a[8:]
        i = b[5:7]
        j = b[8:]
        return f"{x}/{y}-{i}/{j}"

    def get_Fill(self, hexcode):
        return PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")

    def run(self):
        df1 = sorted(self.df1, key=lambda k: k["agent_pk"])

        data = {}
        for item in df1:
            supervisor = item["supervisor_name"]
            collector = item["agent_name"]
            for detail in item["skill_details"]:
                skill_name = detail["skill_name"]
                score = detail["skill_score"]
                call_count = detail["call_count"]
                key = (id, collector, skill_name)
                if key not in data:
                    data[key] = {
                        "supervisor_name": supervisor,
                        "collector_name": collector,
                        "skill_name": skill_name,
                        "call_count": call_count,
                        f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}": 0,
                    }
                if item in df1:
                    data[key][
                        f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}"
                    ] = score

        result = []
        result.append(
            [
                "Supervisor",
                "Collector",
                "Skill",
                "Total Calls",
                f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}",
            ]
        )
        for item in data.values():
            row = [
                item["supervisor_name"],
                item["collector_name"],
                item["skill_name"],
                item["call_count"],
                item[
                    f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}"
                ],
            ]
            result.append(row)

        rule_dict = {d["ruleId"]: d["display_name"] for d in self.rule_desc}

        for lst in result:
            try:
                lst[2] = rule_dict[lst[2]]
            except KeyError:
                continue

        new_df = pd.DataFrame(result)

        new_df.to_excel(
            f"Junk/excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx",
            index=False,
            header=False,
        )

        read_saved_df = pd.read_excel(
            f"Junk/excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx",
        )

        required = [
            "Facts on the table: You have a total amount due of ___. Can you pay that today?",
            "When can you pay the remaining balance of $___?",
            "Diagnostic: How much can you pay today?",
            "Sell Rapid Payment",
            "Did we ask for the Reason for Delinquency",
            "Create Urgency",
            "Recap Expectations of Customer",
        ]

        elem_df = []

        for row in read_saved_df.iterrows():
            if row[1]["Skill"] in required:
                elem_df.append(row[1])

        df = pd.DataFrame(elem_df)

        grouped_df = df.groupby(["Supervisor", "Collector", "Total Calls"])
        result_df = grouped_df.apply(
            lambda x: x.sort_values(
                by="Skill",
                key=lambda col: col.map({k: i for i, k in enumerate(required)}),
            )
        )
        result_df = result_df.reset_index(drop=True)
        result_df.to_excel(
            f"Junk/elem_excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx",
            index=False,
        )

        wb = openpyxl.load_workbook(
            f"Junk/elem_excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx"
        )

        # Select the worksheet to work with
        ws = wb.active

        # Apply conditional formatting based on the values in the "Score 1" column
        for row in ws.iter_rows(min_row=1, min_col=5, max_col=5):
            for cell in row:
                if (
                    cell.value
                    == self.date_converter(
                        self.first_file_start_date, self.first_file_end_date
                    )
                    or cell.value == None
                ):
                    continue
                if 0.0 <= float(cell.value) <= 0.9:
                    cell.fill = self.red_fill
                elif 1.0 <= float(cell.value) <= 1.9:
                    cell.fill = self.yellow_fill
                elif float(cell.value) == 2.0:
                    cell.fill = self.green_fill

        # Set border style
        border = styles.Side(style="thin")
        border_style = styles.Border(
            left=border, right=border, top=border, bottom=border
        )

        # Set font style
        font = styles.Font(size=14)

        # Apply styling to rows and columns with data
        for sheet in wb:
            # Set column width based on maximum content length
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column name
                for cell in column_cells:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                adjusted_width = (
                    max_length + 2
                ) * 1.2  # Adjusting width to accommodate the content better
                sheet.column_dimensions[column].width = adjusted_width

            # Apply border and font style to rows
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = border_style
                    cell.font = font

        # Save the modified file
        wb.save(
            f"weekly_trends_report/excel_weekly_call_analyzer_result_from_({self.first_file_start_date})-({self.first_file_end_date}).xlsx"
        )
        print("\nTrends Generated!\nCheck ```weekly_trends_report``` Folder.\n")


ConverResultToExcel().run()
