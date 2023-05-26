import pandas as pd
import json
import openpyxl
from openpyxl.styles import PatternFill
import openpyxl.styles as styles


class ConverResultToExcel(object):
    def __init__(self) -> None:
        self.first_file_start_date = input("First JSON File Start Date : ")
        self.first_file_end_date = input("First JSON File End Date : ")

        self.df1 = self.load_json(self.first_file_start_date, self.first_file_end_date)

        self.yellow_fill = self.get_Fill("FFF6BD")
        self.red_fill = self.get_Fill("FF9F9F")
        self.green_fill = self.get_Fill("CEEDC7")

        self.rule_desc = self.load_rules()

    def load_json(self, date_start: str, date_end: str) -> dict:
        while True:
            try:
                return json.load(
                    open(
                        input(
                            f"Enter The Path For {date_start} to {date_end} Result : "
                        ),
                        "r",
                    )
                )
            except FileNotFoundError:
                print("\nNot A Valid Path!\n")

    def load_rules(self) -> list:
        return [
            {
                "ruleId": "rules.open.confirm_customer_id.score",
                "display_name": "ID Customer by First and Last Name (Including Suffix)",
                "rule_description": "Agent must obtain from the customer BOTH first and last name.  This may be in a single statement from the customer or multiple statements",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.confirm_agent_id.score",
                "display_name": "ID Self by First and Last Name and Company",
                "rule_description": "Agent must open the call by stating their first and last name including Ally Financial",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.develop_rpc.score",
                "display_name": "Develop RPC Info",
                "rule_description": "Agent must gain rpc info if customer not available (number, work, time available, etc.)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.outbound.state_call_monitor.score",
                "display_name": "Provide Recording Disclosure",
                "rule_description": 'Agent must state that the call may be "monitored" and/or "recorded" for quality assurance purposes.',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.state_mini_miranda.score",
                "display_name": "Provide Mini-Miranda (State Specifics)",
                "rule_description": 'If customer\'s address is in certain states, agent must include the statement "This is an attempt to collect a debt and any information obtained will be used for that purpose."',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.outbound.call_purpose.score",
                "display_name": "State the Purpose of your Call",
                "rule_description": "Agent must provide the Year, Make and Model of the vehicle they are calling the customer about",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.pause_listen.score",
                "display_name": "Pause / Listen",
                "description": "If customer makes an emotional statement, the agent should acknowledge the emotion and statement. They should not include any questions or attempts to hit the next step",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.inbound.account_id.score",
                "display_name": "Ask for Account Number or SS#",
                "rule_description": "Agent must obtain EITHER the account number or SSN to locate the account.  They can ask for both or just one.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.open.inbound.verify_identity.score",
                "display_name": "Ask for the Customer ID",
                "rule_description": "Agent must obtain 2 pieces of verification information.  This can be any of the following: Date of Birth, Last 4 of SSN (if SSN NOT used to pull up the account in item 9), Address,",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.facts_on_the_table.pay_total_amount_due_today.score",
                "display_name": "Facts on the table: You have a total amount due of ___. Can you pay that today?",
                "rule_description": 'This must be stated verbatim.  Agent should include "total amount due <$amount> " and "pay" "today"',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.dqs.how_much_pay_today.score",
                "display_name": "Diagnostic: How much can you pay today?",
                "rule_description": "Applicable if the answer to Facts on the Table is NO.  Question should be asked before asking when the customer can make a payment.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.dqs.when_pay_remaining.score",
                "display_name": "When can you pay the remaining balance of $___?",
                "rule_description": "Applicable if the answer to Facts on the Table is NO.  Question should be asked after asking how much the customer can pay today and BEFORE the agent asks additional questions.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.paying.gain_tad.score",
                "display_name": "Gained Initial Balance Commitment",
                "rule_description": "Agent obtains commitment for full balance in 2 installments prior to next due date.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.rapid_payment.score",
                "display_name": "Sell Rapid Payment",
                "rule_description": "Agent asks customer if they want to set up payment online and provides them benefits such as: easy, instant, free processs",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.paying.straight_to_close.score",
                "display_name": "Proceed Straight to Close",
                "rule_description": "Agent does not ask any probing questions regarding income or reason for delinquency and move to skills - Close Solution",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.willing.two.payments.score",
                "display_name": "Gained Initial Balance Commitment",
                "rule_description": "Agent obtains commitment for full balance in 2 installments prior to next due date.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.willing.raise_offer.score",
                "display_name": "Raised Initial Payment Offer",
                "rule_description": "Agent provided a benefit increasing the initial commitment made by the customer",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.vague.partial_vague.score",
                "display_name": "Agreed for Partial/Initial Payment",
                "rule_description": "Agent obtains commitment for 1 payment but customer would not commit to second payment",
                "sectionOffsetStart": "00:39.4",
                "sectionOffsetEnd": "00:57.8",
            },
            {
                "ruleId": "rules.negotiation_flow.probing_questions.score",
                "display_name": "Set the Stage",
                "rule_description": "Agents ask for permission to ask additional questions to further understand the customer's situation ",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.vague.3questions.score",
                "display_name": "Determined Next Income Period & Gained Commitment for Balance or Some Part",
                "rule_description": "Agent asks for when customer will be receiving next income. Agent asks for full remaining amount due on next income date. Agent asks how much customer can pay on next income date (if applicable)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.vague.repeat_until_full.score",
                "display_name": "Attempted to Continue Using Income Until Full Due is Committed/?",
                "rule_description": "Agent repeats income questions regarding pay periods and gaining partial commitments until total amount is committed",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.other_income.score",
                "display_name": "Determined Other Income in the Household",
                "rule_description": "Agent asks if customer is receiving severance (if applicable) and/or if there are alternate sources of income such as savings, 401k, other assets and/or the potential to borrow money from a friend or family member. Agent asks if customer can borrow money from a friend or family member",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.rfd.score",
                "display_name": "Did we ask for the Reason for Delinquency",
                "rule_description": "Agent asks for the reason the account fell behind ",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.sources_income.score",
                "display_name": "Asked for Regular Sources of Income",
                "rule_description": "Agent asks customer about their regular sources of income such as from a job or unemployment.  Agent asks about the frequency of income received.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.create_solution.score",
                "display_name": "Create & Sell Solutions Based on Sources Found",
                "rule_description": "Agent provides solutions such as borrowing money from the friends or family and paying back",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.modification.score",
                "display_name": "PROGRAM THAT FITS",
                "rule_description": "Offered Properly? (Extension, Modification, & Due Date Change)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.negotiation_flow.unwilling.intent.score",
                "display_name": "Determined Customer's Intent (If Applicable / No Solution)",
                "rule_description": "Agents clarifies plan for payment in the future.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.demographics.score",
                "display_name": "Update Demographics",
                "rule_description": "Agent asks for ALL of the following: Customer's address, Garaging location of the vehicle, Best contact phone number, Employer name (if applicable), Employer phone number (if applicable). Agent should not provide any information on file, agent should ask the customer to provide the information.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.urgency.score",
                "display_name": "Create Urgency",
                "rule_description": "State importance of following through with plan for obtaining funds AND recaps negative result of not following through (to avoid additional fees) OR positive result (to get your account current again)",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.partial.remainder.score",
                "display_name": "Create a Mission for the Remainder",
                "rule_description": 'Agent recaps customer\'s plan to obtain remainder of payment and commitment to call back with result.  Recap must include plan/action committed to (ie "talk to your brother about borrowing <remaining amount>") AND date to call back AND time to call back',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.recap.score",
                "display_name": "Recap Expectations of Customer",
                "rule_description": "If in repo status-inform customer car is out for repo/month end consequences, if applicable based on status of account",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.close.thank.score",
                "display_name": "Thanked the Customer for their Time",
                "rule_description": 'Agent expresses gratitude by "thank you" OR "glad" and wishes good day',
                "sectionOffsetStart": "03:08.3",
                "sectionOffsetEnd": "03:17.2",
            },
            {
                "ruleId": "rules.close.none.mission.score",
                "display_name": "Create a Mission",
                "rule_description": 'Agent recaps the customer\'s plan to obtain funds for payment and their commitment to call back with result.  Recap should include plan/action (ie "talk to your brother about borrowing <amount>") AND date to call back AND time to call back',
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.aet.acknowledge.score",
                "display_name": "Acknowledged What was Said",
                "rule_description": "If customer makes an emotional statement, the agent should acknowledge the emotion and statement.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.aet.wiifm.score",
                "display_name": "Transitioned with a WIIFM",
                "rule_description": "Agent expresses desire to assist customer and provides a benefit to the customer for making a payment and/or staying on the line with the agent.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.active_listening.score",
                "display_name": "Active Listening",
                "rule_description": "Agent reflects backs information provided by the customer.  ",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.acknowledge_emotion.score",
                "display_name": "Acknowledged Customer's Emotion",
                "rule_description": "If customer makes an emotional statement, the agent should acknowledge the emotion and statement.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.remove_isolation.score",
                "display_name": "Did the Agent Attempt to Remove the Isolation",
                "rule_description": "Agent expresses desire to help, that they are the right person to assist the customer, lets the customer know there are many others who have experienced what is happening to the customer",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
            {
                "ruleId": "rules.emotional_outburst.bridge.assure_customer.score",
                "display_name": "Transitioned with Assuring the Customer that we can Find a Solution",
                "rule_description": "Agent expresses confidence and asks/tell the customer they wish to work together to find a solution and/or see what options are available.",
                "sectionOffsetStart": "-1",
                "sectionOffsetEnd": "-1",
            },
        ]

    def date_converter(self, p: str, q: str) -> str:
        a = p
        b = q
        x = a[5:7]
        y = a[8:]
        i = b[5:7]
        j = b[8:]
        return f"{x}/{y}-{i}/{j}"

    def get_Fill(self, hexcode):
        return PatternFill(start_color=hexcode, end_color=hexcode, fill_type="solid")

    def run(self):
        df1 = sorted(self.df1, key=lambda k: k["agent_pk"])

        data = {}
        for item in df1:
            supervisor = item["supervisor_name"]
            collector = item["agent_name"]
            for detail in item["skill_details"]:
                skill_name = detail["skill_name"]
                score = detail["skill_score"]
                call_count = detail["call_count"]
                key = (id, collector, skill_name)
                if key not in data:
                    data[key] = {
                        "supervisor_name": supervisor,
                        "collector_name": collector,
                        "skill_name": skill_name,
                        "call_count": call_count,
                        f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}": 0,
                    }
                if item in df1:
                    data[key][
                        f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}"
                    ] = score

        result = []
        result.append(
            [
                "Supervisor",
                "Collector",
                "Skill",
                "Total Calls",
                f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}",
            ]
        )
        for item in data.values():
            row = [
                item["supervisor_name"],
                item["collector_name"],
                item["skill_name"],
                item["call_count"],
                item[
                    f"{self.date_converter(self.first_file_start_date,self.first_file_end_date)}"
                ],
            ]
            result.append(row)

        rule_dict = {d["ruleId"]: d["display_name"] for d in self.rule_desc}

        for lst in result:
            try:
                lst[2] = rule_dict[lst[2]]
            except KeyError:
                continue

        new_df = pd.DataFrame(result)

        new_df.to_excel(
            f"NoColorFinalResult/excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx",
            index=False,
            header=False,
        )

        read_saved_df = pd.read_excel(
            f"NoColorFinalResult/excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx",
        )

        required = [
            "Facts on the table: You have a total amount due of ___. Can you pay that today?",
            "When can you pay the remaining balance of $___?",
            "Diagnostic: How much can you pay today?",
            "Sell Rapid Payment",
            "Did we ask for the Reason for Delinquency",
            "Create Urgency",
            "Recap Expectations of Customer",
        ]

        elem_df = []

        for row in read_saved_df.iterrows():
            if row[1]["Skill"] in required:
                elem_df.append(row[1])

        df = pd.DataFrame(elem_df)

        grouped_df = df.groupby(["Supervisor", "Collector", "Total Calls"])
        result_df = grouped_df.apply(
            lambda x: x.sort_values(
                by="Skill",
                key=lambda col: col.map({k: i for i, k in enumerate(required)}),
            )
        )
        result_df = result_df.reset_index(drop=True)
        result_df.to_excel(
            f"NoColorFinalResult/elem_excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx",
            index=False,
        )

        wb = openpyxl.load_workbook(
            f"NoColorFinalResult/elem_excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx"
        )

        # Select the worksheet to work with
        ws = wb.active

        # Apply conditional formatting based on the values in the "Score 1" column
        for row in ws.iter_rows(min_row=1, min_col=5, max_col=5):
            for cell in row:
                if (
                    cell.value
                    == self.date_converter(
                        self.first_file_start_date, self.first_file_end_date
                    )
                    or cell.value == None
                ):
                    continue
                if 0.0 <= float(cell.value) <= 0.9:
                    cell.fill = self.red_fill
                elif 1.0 <= float(cell.value) <= 1.9:
                    cell.fill = self.yellow_fill
                elif float(cell.value) == 2.0:
                    cell.fill = self.green_fill

        # Set border style
        border = styles.Side(style="thin")
        border_style = styles.Border(
            left=border, right=border, top=border, bottom=border
        )

        # Set font style
        font = styles.Font(size=14)

        # Apply styling to rows and columns with data
        for sheet in wb:
            # Set column width based on maximum content length
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column name
                for cell in column_cells:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                adjusted_width = (
                    max_length + 2
                ) * 1.2  # Adjusting width to accommodate the content better
                sheet.column_dimensions[column].width = adjusted_width

            # Apply border and font style to rows
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = border_style
                    cell.font = font

        # Save the modified file
        wb.save(
            f"Result/excel_weekly_call_analyzer_result_form_({self.first_file_start_date})-({self.first_file_end_date}).xlsx"
        )


ConverResultToExcel().run()
